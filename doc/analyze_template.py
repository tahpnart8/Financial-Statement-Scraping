import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('DMC model.xlsx', data_only=True)
print(f"Sheet names: {wb.sheetnames}")

for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n{'='*80}")
    print(f"Sheet: {s}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Print first 80 rows to understand structure
    print(f"\n--- Content (first 80 rows) ---")
    for row in ws.iter_rows(min_row=1, max_row=min(80, ws.max_row), max_col=min(20, ws.max_column), values_only=False):
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                row_data.append(f"{cell.coordinate}: {val}")
        if row_data:
            print(" | ".join(row_data))
