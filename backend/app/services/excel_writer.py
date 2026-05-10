"""
Excel Writer Service
====================
Dynamically generates an Excel workbook populated with financial data.
The template is generated programmatically (not from a static .xlsx file)
to support dynamic year ranges.
"""
from __future__ import annotations

import logging
import os
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from app.services.data_mapper import PL_ROW_MAP, BS_ROW_MAP, CF_ROW_MAP

logger = logging.getLogger(__name__)


# ── Style Constants ────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Times New Roman", size=11, bold=True)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Times New Roman", size=10)
NUMBER_FORMAT = '#,##0'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


# ── Vietnamese Labels ──────────────────────────────────────────────────────────

PL_LABELS = {
    3: "1. Doanh thu bán hàng và cung cấp dịch vụ",
    4: "2. Các khoản giảm trừ doanh thu",
    5: "3. Doanh thu thuần về bán hàng và cung cấp dịch vụ",
    6: "4. Giá vốn hàng bán",
    7: "5. Lợi nhuận gộp về bán hàng và cung cấp dịch vụ",
    8: "6. Doanh thu hoạt động tài chính",
    9: "7. Chi phí tài chính",
    10: "   Trong đó: Chi phí lãi vay",
    11: "8. Phần lãi/lỗ trong công ty liên doanh, liên kết",
    12: "9. Chi phí bán hàng",
    13: "10. Chi phí quản lý doanh nghiệp",
    14: "11. Lợi nhuận thuần từ hoạt động kinh doanh",
    15: "12. Thu nhập khác",
    16: "13. Chi phí khác",
    17: "14. Lợi nhuận khác",
    18: "15. Tổng lợi nhuận kế toán trước thuế",
    19: "16. Chi phí thuế TNDN hiện hành",
    20: "Thuế suất",
    21: "17. Chi phí thuế TNDN hoãn lại",
    22: "18. Lợi nhuận sau thuế thu nhập doanh nghiệp",
    23: "Lợi ích của cổ đông thiểu số",
    24: "% LNST",
    25: "Lợi nhuận sau thuế của cổ đông của Công ty mẹ",
    27: "19. Lãi cơ bản trên cổ phiếu (*) (VNĐ)",
}

BS_LABELS = {
    32: "TÀI SẢN",
    33: "A. TÀI SẢN NGẮN HẠN",
    34: "I. Tiền và các khoản tương đương tiền",
    35: "1. Tiền",
    36: "2. Các khoản tương đương tiền",
    37: "II. Đầu tư tài chính ngắn hạn",
    38: "1. Chứng khoán kinh doanh",
    39: "2. Dự phòng giảm giá chứng khoán kinh doanh (*)",
    40: "3. Đầu tư nắm giữ đến ngày đáo hạn",
    41: "III. Các khoản phải thu ngắn hạn",
    42: "1. Phải thu ngắn hạn của khách hàng",
    43: "2. Trả trước cho người bán ngắn hạn",
    44: "3. Phải thu nội bộ ngắn hạn",
    45: "4. Phải thu theo tiến độ kế hoạch hợp đồng xây dựng",
    46: "5. Phải thu về cho vay ngắn hạn",
    47: "6. Phải thu ngắn hạn khác",
    48: "7. Dự phòng phải thu ngắn hạn khó đòi (*)",
    49: "8. Tài sản thiếu chờ xử lý",
    50: "IV. Hàng tồn kho",
    51: "1. Hàng tồn kho",
    52: "2. Dự phòng giảm giá hàng tồn kho (*)",
    53: "V. Tài sản ngắn hạn khác",
    54: "1. Chi phí trả trước ngắn hạn",
    55: "2. Thuế GTGT được khấu trừ",
    56: "3. Thuế và các khoản khác phải thu của nhà nước",
    57: "4. Giao dịch mua bán lại trái phiếu chính phủ",
    58: "5. Tài sản ngắn hạn khác",
    60: "B. TÀI SẢN DÀI HẠN",
    61: "I. Các khoản phải thu dài hạn",
    62: "1. Phải thu dài hạn của khách hàng",
    63: "2. Trả trước cho người bán dài hạn",
    64: "3. Vốn kinh doanh ở các đơn vị trực thuộc",
    65: "4. Phải thu nội bộ dài hạn",
    66: "5. Phải thu về cho vay dài hạn",
    67: "6. Phải thu dài hạn khác",
    68: "7. Dự phòng phải thu dài hạn khó đòi (*)",
    69: "II. Tài sản cố định",
    70: "1. Tài sản cố định hữu hình",
    71: "      - Nguyên giá",
    72: "      - Giá trị hao mòn lũy kế (*)",
    73: "2. Tài sản cố định thuê tài chính",
    74: "      - Nguyên giá",
    75: "      - Giá trị hao mòn lũy kế (*)",
    76: "3. Tài sản cố định vô hình",
    77: "      - Nguyên giá",
    78: "      - Giá trị hao mòn lũy kế (*)",
    79: "III. Bất động sản đầu tư",
    80: "      - Nguyên giá",
    81: "      - Giá trị hao mòn lũy kế (*)",
    82: "IV. Tài sản dở dang dài hạn",
    83: "1. Chi phí sản xuất, kinh doanh dở dang dài hạn",
    84: "2. Chi phí xây dựng cơ bản dở dang",
    85: "V. Đầu tư tài chính dài hạn",
    86: "1. Đầu tư vào công ty con",
    87: "2. Đầu tư vào công ty liên kết, liên doanh",
    88: "3. Đầu tư góp vốn vào đơn vị khác",
    89: "4. Dự phòng đầu tư tài chính dài hạn (*)",
    90: "5. Đầu tư nắm giữ đến ngày đáo hạn",
    91: "6. Đầu tư dài hạn khác",
    92: "VI. Tài sản dài hạn khác",
    93: "1. Chi phí trả trước dài hạn",
    94: "2. Tài sản thuế thu nhập hoãn lại",
    95: "3. Thiết bị, vật tư, phụ tùng thay thế dài hạn",
    96: "4. Tài sản dài hạn khác",
    97: "VII. Lợi thế thương mại",
    98: "TỔNG CỘNG TÀI SẢN",
    100: "NGUỒN VỐN",
    101: "A. NỢ PHẢI TRẢ",
    102: "I. Nợ ngắn hạn",
    103: "1. Phải trả người bán ngắn hạn",
    104: "2. Người mua trả tiền trước ngắn hạn",
    105: "3. Thuế và các khoản phải nộp Nhà nước",
    106: "4. Phải trả người lao động",
    107: "5. Chi phí phải trả ngắn hạn",
    108: "6. Phải trả nội bộ ngắn hạn",
    109: "7. Phải trả theo tiến độ kế hoạch hợp đồng xây dựng",
    110: "8. Doanh thu chưa thực hiện ngắn hạn",
    111: "9. Phải trả ngắn hạn khác",
    112: "10. Vay và nợ thuê tài chính ngắn hạn",
    113: "11. Dự phòng phải trả ngắn hạn",
    114: "12. Quỹ khen thưởng, phúc lợi",
    115: "13. Quỹ bình ổn giá",
    116: "14. Giao dịch mua bán lại trái phiếu Chính phủ",
    117: "II. Nợ dài hạn",
    118: "1. Phải trả người bán dài hạn",
    119: "2. Người mua trả tiền trước dài hạn",
    120: "3. Chi phí phải trả dài hạn",
    121: "4. Phải trả nội bộ về vốn kinh doanh",
    122: "5. Phải trả nội bộ dài hạn",
    123: "6. Doanh thu chưa thực hiện dài hạn",
    124: "7. Phải trả dài hạn khác",
    125: "8. Vay và nợ thuê tài chính dài hạn",
    126: "9. Trái phiếu chuyển đổi",
    127: "10. Cổ phiếu ưu đãi (Nợ)",
    128: "11. Thuế thu nhập hoãn lại phải trả",
    129: "12. Dự phòng phải trả dài hạn",
    130: "13. Quỹ phát triển khoa học và công nghệ",
    131: "14. Dự phòng trợ cấp mất việc làm",
    132: "B. VỐN CHỦ SỞ HỮU",
    133: "I. Vốn chủ sở hữu",
    134: "1. Vốn góp của chủ sở hữu",
    135: "      - Cổ phiếu phổ thông có quyền biểu quyết",
    136: "      - Cổ phiếu ưu đãi",
    137: "2. Thặng dư vốn cổ phần",
    138: "3. Quyền chọn chuyển đổi trái phiếu",
    139: "4. Vốn khác của chủ sở hữu",
    140: "5. Cổ phiếu quỹ (*)",
    141: "6. Chênh lệch đánh giá lại tài sản",
    142: "7. Chênh lệch tỷ giá hối đoái",
    143: "8. Quỹ đầu tư phát triển",
    144: "9. Quỹ hỗ trợ sắp xếp doanh nghiệp",
    145: "10. Quỹ khác thuộc vốn chủ sở hữu",
    146: "11. Lợi nhuận sau thuế chưa phân phối",
    147: "      - LNST chưa phân phối lũy kế đến cuối kỳ trước",
    148: "      - LNST chưa phân phối kỳ này",
    149: "12. Nguồn vốn đầu tư XDCB",
    150: "13. Lợi ích cổ đông không kiểm soát",
    151: "14. Quỹ dự phòng tài chính",
    152: "II. Nguồn kinh phí và quỹ khác",
    153: "1. Nguồn kinh phí",
    154: "2. Nguồn kinh phí đã hình thành TSCĐ",
    155: "C. LỢI ÍCH CỔ ĐÔNG THIỂU SỐ",
    156: "TỔNG CỘNG NGUỒN VỐN",
}

CF_LABELS = {
    160: "I. Lưu chuyển tiền từ hoạt động kinh doanh",
    161: "Lãi lỗ sau thuế",
    162: "Khấu hao TSCĐ",
    163: "Thay đổi vốn lưu động",
    164: "Thay đổi khác",
    165: "Lưu chuyển tiền thuần từ hoạt động kinh doanh",
    167: "II. Lưu chuyển tiền từ hoạt động đầu tư",
    168: "1. Tiền chi để mua sắm, xây dựng TSCĐ và các tài sản dài hạn khác",
    169: "2. Tiền thu từ thanh lý, nhượng bán TSCĐ và các tài sản dài hạn khác",
    170: "3. Tiền chi cho vay, mua các công cụ nợ của đơn vị khác",
    171: "4. Tiền thu hồi cho vay, bán lại các công cụ nợ của đơn vị khác",
    172: "5. Tiền chi đầu tư góp vốn vào đơn vị khác",
    173: "6. Tiền thu hồi đầu tư góp vốn vào đơn vị khác",
    174: "7. Tiền thu lãi cho vay, cổ tức và lợi nhuận được chia",
    175: "8. Tăng giảm tiền gửi ngân hàng có kỳ hạn",
    176: "9. Mua lại khoản góp vốn của cổ đông thiểu số trong công ty con",
    177: "10. Tiền thu khác từ hoạt động đầu tư",
    178: "11. Tiền chi khác cho hoạt động đầu tư",
    179: "Lưu chuyển tiền thuần từ hoạt động đầu tư",
    181: "III. Lưu chuyển tiền từ hoạt động tài chính",
    182: "1. Tiền thu từ phát hành cổ phiếu, nhận vốn góp của chủ sở hữu",
    183: "2. Tiền chi trả vốn góp cho các chủ sở hữu, mua lại cổ phiếu của doanh nghiệp đã phát hành",
    184: "3. Tiền thu từ đi vay",
    185: "4. Tiền trả nợ gốc vay",
    186: "5. Tiền trả nợ gốc thuê tài chính",
    187: "6. Cổ tức, lợi nhuận đã trả cho chủ sở hữu",
    188: "7. Tiền thu khác từ hoạt động tài chính",
    189: "8. Tiền chi khác cho hoạt động tài chính",
    190: "Lưu chuyển tiền thuần từ hoạt động tài chính",
    192: "Lưu chuyển tiền thuần trong kỳ",
    193: "Tiền và tương đương tiền đầu kỳ",
    194: "Ảnh hưởng của thay đổi tỷ giá hối đoái quy đổi ngoại tệ",
    195: "Tiền và tương đương tiền cuối kỳ",
}

REVERSE_PL_MAP = {v: k for k, v in PL_ROW_MAP.items()}
REVERSE_BS_MAP = {v: k for k, v in BS_ROW_MAP.items()}
REVERSE_CF_MAP = {v: k for k, v in CF_ROW_MAP.items()}

def _apply_cell_style(ws, row: int, col: int, value: Any, is_header: bool = False,
                       is_section: bool = False, fmt: str | None = None):
    """Apply value and styling to a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)

    if is_header:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif is_section:
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
    else:
        cell.font = DATA_FONT

    if fmt:
        cell.number_format = fmt


def _write_year_headers(ws, start_row: int, label: str, years: list[int],
                        label_col: int = 2):
    """Write section header + year column headers."""
    _apply_cell_style(ws, start_row, label_col, label, is_header=True)
    for i, year in enumerate(years):
        _apply_cell_style(ws, start_row, label_col + 1 + i, year, is_header=True)


def _write_data_row(ws, row: int, label: str, data: dict[int, Any] | None,
                    years: list[int], label_col: int = 2,
                    is_section: bool = False, fmt: str | None = NUMBER_FORMAT):
    """Write a label + yearly values into a row. If data is None, leave value cells blank."""
    _apply_cell_style(ws, row, label_col, label, is_section=is_section)
    if data is None:
        for i, year in enumerate(years):
            _apply_cell_style(ws, row, label_col + 1 + i, "", is_section=is_section)
    else:
        for i, year in enumerate(years):
            val = data.get(year, 0)
            if pd.isna(val):
                val = 0
            _apply_cell_style(ws, row, label_col + 1 + i, val, fmt=fmt, is_section=is_section)


# ── Sheet Builders ─────────────────────────────────────────────────────────────

def _build_bctc_sheet(
    wb: Workbook,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    years: list[int],
):
    """Build the main 'BCTC du phong' sheet with PL + BS + CF sections."""
    ws = wb.active
    ws.title = "BCTC du phong"

    # Set column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 55
    for i, _ in enumerate(years):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 18

    # ── Profit & Loss Section ──
    _write_year_headers(ws, 1, "PL", years)

    is_data = mapped_data.get("income_statement", {})
    
    # Calculate tax rate mapping manually
    tax_rate_data = {}
    pbt = is_data.get("profit_before_tax", {})
    tax = is_data.get("current_tax_expense", {})
    for y in years:
        pbt_val = pbt.get(y, 0)
        tax_val = tax.get(y, 0)
        if pbt_val and pbt_val != 0:
            tax_rate_data[y] = tax_val / pbt_val
            
    # Calculate Profit Margin (% LNST)
    margin_data = {}
    rev = is_data.get("net_revenue", {})
    pat = is_data.get("profit_after_tax", {})
    for y in years:
        rev_val = rev.get(y, 0)
        pat_val = pat.get(y, 0)
        if rev_val and rev_val != 0:
            margin_data[y] = pat_val / rev_val
            
    for row_idx, label in PL_LABELS.items():
        is_section = "Tổng" in label or "Lợi nhuận" in label or label.isupper()
        fmt = PERCENT_FORMAT if "%" in label or "Thuế suất" in label else NUMBER_FORMAT
        
        if row_idx == 20: # Thuế suất
            _write_data_row(ws, row_idx, label, tax_rate_data, years, is_section=is_section, fmt=fmt)
            continue
        if row_idx == 24: # % LNST
            _write_data_row(ws, row_idx, label, margin_data, years, is_section=is_section, fmt=fmt)
            continue
            
        internal_key = REVERSE_PL_MAP.get(row_idx)
        if not internal_key:
            _write_data_row(ws, row_idx, label, None, years, is_section=is_section)
        else:
            data = is_data.get(internal_key, {})
            _write_data_row(ws, row_idx, label, data, years, is_section=is_section, fmt=fmt)

    # ── Balance Sheet Section ──
    bs_start = 30
    _write_year_headers(ws, bs_start, "BS", years)

    bs_data = mapped_data.get("balance_sheet", {})
    for row_idx, label in BS_LABELS.items():
        actual_row = row_idx
        is_section = label.isupper() or label.startswith(tuple("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        
        internal_key = REVERSE_BS_MAP.get(row_idx)
        if not internal_key:
            _write_data_row(ws, actual_row, label, None, years, is_section=is_section)
        else:
            data = bs_data.get(internal_key, {})
            _write_data_row(ws, actual_row, label, data, years, is_section=is_section)

    # ── Cash Flow Section ──
    cf_start = 158
    _write_year_headers(ws, cf_start, "CF", years)

    cf_data = mapped_data.get("cash_flow", {})
    for row_idx, label in CF_LABELS.items():
        actual_row = row_idx
        is_section = label.isupper() or label.startswith(tuple("ABCDEFGHIJKLMNOPQRSTUVWXYZ")) or "Lưu chuyển tiền thuần" in label
        
        internal_key = REVERSE_CF_MAP.get(row_idx)
        if not internal_key:
            _write_data_row(ws, actual_row, label, None, years, is_section=is_section)
        else:
            data = cf_data.get(internal_key, {})
            _write_data_row(ws, actual_row, label, data, years, is_section=is_section)


def _build_revenue_sheet(
    wb: Workbook,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    years: list[int],
):
    """Build the 'Doanh thu' (Revenue) analysis sheet."""
    ws = wb.create_sheet("Doanh thu")
    ws.column_dimensions["A"].width = 40
    for i, _ in enumerate(years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    _write_year_headers(ws, 1, "I/ Dự phóng doanh thu", years, label_col=1)

    is_data = mapped_data.get("income_statement", {})

    # Net revenue
    _write_data_row(ws, 3, "Doanh thu thuần", is_data.get("net_revenue", {}),
                    years, label_col=1, is_section=True)

    # Revenue growth rate
    net_rev = is_data.get("net_revenue", {})
    growth = {}
    sorted_years = sorted(years)
    for i in range(1, len(sorted_years)):
        prev = net_rev.get(sorted_years[i - 1], 0)
        curr = net_rev.get(sorted_years[i], 0)
        if prev and prev != 0:
            growth[sorted_years[i]] = (curr - prev) / prev
    _write_data_row(ws, 4, "% tăng trưởng", growth, years, label_col=1,
                    fmt=PERCENT_FORMAT)

    # COGS
    row = 6
    _apply_cell_style(ws, row, 1, "II/ Dự phóng giá vốn", is_section=True)
    row += 1
    _write_data_row(ws, row, "Giá vốn", is_data.get("cost_of_goods_sold", {}),
                    years, label_col=1)

    # COGS/Revenue ratio
    cogs = is_data.get("cost_of_goods_sold", {})
    cogs_ratio = {}
    for y in years:
        rev = net_rev.get(y, 0)
        c = cogs.get(y, 0)
        if rev and rev != 0:
            cogs_ratio[y] = c / rev
    _write_data_row(ws, row + 1, "% doanh thu thuần", cogs_ratio, years,
                    label_col=1, fmt=PERCENT_FORMAT)

    # Selling expenses
    row += 3
    _write_data_row(ws, row, "CP bán hàng", is_data.get("selling_expenses", {}),
                    years, label_col=1)
    sell_exp = is_data.get("selling_expenses", {})
    sell_ratio = {y: sell_exp.get(y, 0) / net_rev.get(y, 1) for y in years if net_rev.get(y, 0)}
    _write_data_row(ws, row + 1, "% doanh thu", sell_ratio, years,
                    label_col=1, fmt=PERCENT_FORMAT)

    # Admin expenses
    row += 3
    _write_data_row(ws, row, "CP quản lý DN", is_data.get("admin_expenses", {}),
                    years, label_col=1)
    admin_exp = is_data.get("admin_expenses", {})
    admin_ratio = {y: admin_exp.get(y, 0) / net_rev.get(y, 1) for y in years if net_rev.get(y, 0)}
    _write_data_row(ws, row + 1, "% doanh thu", admin_ratio, years,
                    label_col=1, fmt=PERCENT_FORMAT)


def _build_working_capital_sheet(
    wb: Workbook,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    years: list[int],
):
    """Build the 'Vốn lưu động' (Working Capital) sheet."""
    ws = wb.create_sheet("Vốn lưu động")
    ws.column_dimensions["A"].width = 45
    for i, _ in enumerate(years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    _write_year_headers(ws, 1, "TS ngắn hạn", years, label_col=1)

    bs_data = mapped_data.get("balance_sheet", {})
    is_data = mapped_data.get("income_statement", {})
    net_rev = is_data.get("net_revenue", {})
    cogs = is_data.get("cost_of_goods_sold", {})

    # Short-term receivables section
    row = 3
    items_receivables = [
        ("short_term_receivables", "III. Các khoản phải thu ngắn hạn", True),
        ("trade_receivables", "1. Phải thu ngắn hạn của khách hàng", False),
        ("prepayments_to_suppliers", "2. Trả trước cho người bán ngắn hạn", False),
        ("other_receivables_st", "6. Phải thu ngắn hạn khác", False),
        ("bad_debt_provision_st", "7. Dự phòng phải thu ngắn hạn khó đòi (*)", False),
    ]

    for key, label, is_sec in items_receivables:
        _write_data_row(ws, row, label, bs_data.get(key, {}), years,
                        label_col=1, is_section=is_sec)
        # Add turnover ratio for trade receivables
        if key == "trade_receivables":
            row += 1
            ar = bs_data.get(key, {})
            turnover = {}
            for y in years:
                ar_val = ar.get(y, 0)
                if ar_val and ar_val != 0:
                    turnover[y] = net_rev.get(y, 0) / ar_val
            _write_data_row(ws, row, "Vòng quay phải thu", turnover, years,
                            label_col=1, fmt='0.00')
        row += 2

    # Inventory section
    row += 1
    inv_items = [
        ("inventories", "IV. Hàng tồn kho", True),
        ("inventory_goods", "1. Hàng tồn kho", False),
        ("inventory_provision", "2. Dự phòng giảm giá hàng tồn kho (*)", False),
    ]
    for key, label, is_sec in inv_items:
        _write_data_row(ws, row, label, bs_data.get(key, {}), years,
                        label_col=1, is_section=is_sec)
        if key == "inventory_goods":
            row += 1
            inv = bs_data.get(key, {})
            inv_turnover = {}
            for y in years:
                inv_val = inv.get(y, 0)
                if inv_val and inv_val != 0:
                    inv_turnover[y] = cogs.get(y, 0) / inv_val
            _write_data_row(ws, row, "Vòng quay HTK", inv_turnover, years,
                            label_col=1, fmt='0.00')
        row += 1


def _build_fixed_assets_sheet(
    wb: Workbook,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    years: list[int],
):
    """Build the 'TSCĐ' (Fixed Assets) sheet."""
    ws = wb.create_sheet("TSCĐ")
    ws.column_dimensions["A"].width = 45
    for i, _ in enumerate(years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    _write_year_headers(ws, 1, "TSCĐ hữu hình", years, label_col=1)

    bs_data = mapped_data.get("balance_sheet", {})

    items = [
        ("tangible_fixed_assets", "1. Tài sản cố định hữu hình", True),
        ("intangible_fixed_assets", "3. Tài sản cố định vô hình", False),
        ("investment_property", "III. Bất động sản đầu tư", False),
        ("long_term_wip", "IV. Tài sản dở dang dài hạn", False),
    ]

    row = 3
    for key, label, is_sec in items:
        _write_data_row(ws, row, label, bs_data.get(key, {}), years,
                        label_col=1, is_section=is_sec)
        row += 2


def _build_equity_sheet(
    wb: Workbook,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    years: list[int],
):
    """Build the 'Vốn CSH' (Equity) sheet."""
    ws = wb.create_sheet("Vốn CSH")
    ws.column_dimensions["A"].width = 45
    for i, _ in enumerate(years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    _write_year_headers(ws, 1, "Vốn chủ sở hữu", years, label_col=1)

    bs_data = mapped_data.get("balance_sheet", {})

    items = [
        ("owners_equity", "I. Vốn chủ sở hữu", True),
        ("charter_capital", "1. Vốn góp của chủ sở hữu", False),
        ("share_premium", "2. Thặng dư vốn cổ phần", False),
        ("treasury_shares", "5. Cổ phiếu quỹ (*)", False),
        ("investment_development_fund", "8. Quỹ đầu tư phát triển", False),
        ("retained_earnings", "11. Lợi nhuận sau thuế chưa phân phối", False),
        ("non_controlling_interests", "13. Lợi ích cổ đông không kiểm soát", False),
    ]

    row = 3
    for key, label, is_sec in items:
        _write_data_row(ws, row, label, bs_data.get(key, {}), years,
                        label_col=1, is_section=is_sec)
        row += 1


def _build_financial_income_sheet(
    wb: Workbook,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    years: list[int],
):
    """Build the 'DT&CP tài chính' (Financial Income/Expenses) sheet."""
    ws = wb.create_sheet("DT&CP tài chính")
    ws.column_dimensions["A"].width = 45
    for i, _ in enumerate(years):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    _write_year_headers(ws, 1, "DT & CP tài chính", years, label_col=1)

    is_data = mapped_data.get("income_statement", {})
    bs_data = mapped_data.get("balance_sheet", {})

    items = [
        ("financial_income", "DT tài chính", True, "income_statement"),
        ("financial_expenses", "Chi phí tài chính", False, "income_statement"),
        ("interest_expenses", "Lãi vay", False, "income_statement"),
        ("st_borrowings", "Vay ngắn hạn", False, "balance_sheet"),
    ]

    row = 3
    for key, label, is_sec, source in items:
        data_src = is_data if source == "income_statement" else bs_data
        _write_data_row(ws, row, label, data_src.get(key, {}), years,
                        label_col=1, is_section=is_sec)
        row += 1


def _build_valuation_sheet(wb: Workbook, years: list[int]):
    """Build the 'Định giá' (Valuation) template sheet — empty for user to fill."""
    ws = wb.create_sheet("Định giá")
    ws.column_dimensions["A"].width = 40
    _apply_cell_style(ws, 1, 1, "Định giá", is_header=True)
    _apply_cell_style(ws, 3, 1, "Giá theo FCFE", is_section=True)
    _apply_cell_style(ws, 4, 1, "Giá theo FCFF", is_section=True)
    _apply_cell_style(ws, 6, 1, "Giá trị nội tại", is_section=True)
    _apply_cell_style(ws, 7, 1, "Thị giá")
    _apply_cell_style(ws, 8, 1, "UP/DOWN")


def _build_assumptions_sheet(wb: Workbook, years: list[int]):
    """Build the 'Giả định' (Assumptions) template sheet — empty for user to fill."""
    ws = wb.create_sheet("Giả định vs biểu đồ")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 25

    _apply_cell_style(ws, 2, 1, "STT", is_header=True)
    _apply_cell_style(ws, 2, 2, "Giả định", is_header=True)
    _apply_cell_style(ws, 2, 4, "Giá trị năm ngoái", is_header=True)
    _apply_cell_style(ws, 2, 5, "Giá trị hiện tại", is_header=True)
    _apply_cell_style(ws, 2, 6, "Cách tính", is_header=True)
    _apply_cell_style(ws, 2, 7, "Chú thích", is_header=True)


# ── Public API ─────────────────────────────────────────────────────────────────

def generate_excel(
    ticker: str,
    mapped_data: dict[str, dict[str, dict[int, Any]]],
    year_from: int,
    year_to: int,
    output_dir: str,
) -> str:
    """
    Generate a complete financial model Excel workbook.

    Parameters
    ----------
    ticker : str
        Stock ticker (used in filename).
    mapped_data : dict
        Output of data_mapper.map_financial_data().
    year_from, year_to : int
        Year range for column headers.
    output_dir : str
        Directory to save the output file.

    Returns
    -------
    str
        Absolute path to the generated .xlsx file.
    """
    years = list(range(year_from, year_to + 1))

    wb = Workbook()

    # Build ONLY the main BCTC sheet as requested for this version
    _build_bctc_sheet(wb, mapped_data, years)

    # Save
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{ticker}_BCTC_{year_from}_{year_to}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    logger.info(f"Excel file generated: {filepath}")
    return filepath
